"""
Emovr Migration — Step 1: Scan & Clean

Reads messy Excel files from input/ and produces normalized CSVs in cleaned/.
Rules are based on migration.md.

Usage:
    python clean.py
"""

import openpyxl
import csv
import os
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
CLEANED_DIR = BASE_DIR / "cleaned"

# ---------------------------------------------------------------------------
# Normalization rules (from migration.md)
# ---------------------------------------------------------------------------

KLANT_NORMALISATIE = {
    "verwoert": "Restauratiebedrijf Verwoerd",
    "verwoerd": "Restauratiebedrijf Verwoerd",
    "restauratie bedrijf verwoerd": "Restauratiebedrijf Verwoerd",
    "restauratiebedrijf verwoerd": "Restauratiebedrijf Verwoerd",
    "ss teknik as": "SS Teknikk AS",
    "ss teknikk as": "SS Teknikk AS",
    "sloetjes montage": "A Sloetjes montage",
    "a sloetjes montage": "A Sloetjes montage",
    "compraan bv": "Compraan BV",
    "gkb buiteninrichting": "GKB Buiteninrichting",
    "mulder montage bv": "Mulder Montage BV",
    "mulder montage bv ": "Mulder Montage BV",
    "van gelder": "Van Gelder",
    "van gelder ": "Van Gelder",
    "brouwer": "Brouwer",
    "brouwer egalisatie": "Brouwer Egalisatie",  # TODO: klant vragen of dit zelfde bedrijf is als Brouwer
    "tuboma": "Tuboma",
    "emovr": "Emovr BV",
    "emovr bv": "Emovr BV",
    " ss teknikk as": "SS Teknikk AS",
    "b&b diensten": "B&B Diensten",
    "van de ende": "Van de Ende",
}

# Contact name typo corrections
CONTACT_NORMALISATIE = {
    "marinus slinerland": "Marinus Slingerland",
    "marinus slingerland": "Marinus Slingerland",
}

TYPE_NORMALISATIE = {
    "trc 1.5": "TRC1.5",
    "trc1,5": "TRC1.5",
    "trc1.5": "TRC1.5",
    "cr0,5": "CR0.5",
    "cr0.5": "CR0.5",
    "gl1,0": "GL1.0",
    "glb1,2": "GLB1.2",
    "sch1.4": "SCH1.4",
    "du250": "DU250",
    "hm0,5": "HM0.5",
    "hm1,2": "HM1.2",
}

# Serienummer correcties (TODO: invullen na klant-feedback)
# Format: (original_sn, klant) → corrected_sn
SERIENUMMER_KLANT_CORRECTIES = {
    # ("EMTRC15092500", "SS Teknikk AS"): "EMTRC1509250007",  # BLOCKED: aan klant vragen
    # ("EMTRC15092500", ""): "EMTRC1509250010",               # BLOCKED: aan klant vragen
}

# Machine-to-klant corrections (TODO: invullen na klant-feedback)
MACHINE_KLANT_CORRECTIES = {
    # "EMTRC150925009": "???",  # BLOCKED: aan klant vragen welke klant
}

# Sheets to skip entirely
SKIP_SHEETS = {"rongen systeem", "brouwer egalisatie"}

# Klant-tabs (per-customer sheets)
KLANT_TABS = {
    "Compraan BV", "GKB Buiteninrichting", "Mulder montage BV",
    "Restauratiebedrijf Verwoerd", "A Sloetjes montage", "SS Teknikk AS",
    "Van Gelder", "Brouwer", "Tuboma",
}

# Productie-overzicht sheets (aanbouwdelen)
AANBOUWDEEL_SHEETS = {"Dumper optie", "Crane optie", "Glasbok", "Hefmast", "Loader"}

# Master machine sheet
MACHINE_SHEET = "machines "


def normalize_klant(name):
    """Normalize customer name using lookup table."""
    if name is None:
        return ""
    cleaned = str(name).strip()
    if not cleaned:
        return ""
    key = cleaned.lower().strip()
    return KLANT_NORMALISATIE.get(key, cleaned)


def normalize_type(type_str):
    """Normalize machine/part type code."""
    if type_str is None:
        return ""
    cleaned = str(type_str).strip()
    if not cleaned:
        return ""
    key = cleaned.lower().strip()
    return TYPE_NORMALISATIE.get(key, cleaned)


def normalize_value(val):
    """Handle N.V.T, ?, and other placeholder values."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.upper() in ("N.V.T", "N,V,T", "N.V.T.", "NVT", "?", ""):
        return ""
    return s


def format_date(val):
    """Format datetime to ISO date string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s.upper() in ("N.V.T", "N,V,T", "?", ""):
        return ""
    return s


def is_header_row(row_values):
    """Detect if a row is a repeated header row."""
    if not row_values:
        return False
    first = str(row_values[0]).strip().lower() if row_values[0] else ""
    return first in ("klant/bedrijf", "aanbouwdeel", "machine")


def is_data_row(row_values):
    """Check if a row contains actual data (not empty, not just text)."""
    non_empty = [v for v in row_values if v is not None and str(v).strip()]
    if len(non_empty) < 2:
        return False
    # If first cell is very long text, it's likely a comment
    first = str(row_values[0]).strip() if row_values[0] else ""
    if len(first) > 60 and non_empty == [row_values[0]]:
        return False
    return True


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------

def extract_klanten_and_contacts(wb):
    """Extract unique customers and contact persons from klant-tabs."""
    klanten = {}  # name -> {fields}
    contacten = []  # list of {name, klant}

    for sheet_name in KLANT_TABS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not is_data_row(list(row)):
                continue
            if is_header_row(list(row)):
                continue

            klant_raw = row[0]
            contact_raw = row[5] if len(row) > 5 else None

            if klant_raw and str(klant_raw).strip():
                klant_name = normalize_klant(klant_raw)
                if klant_name and klant_name not in klanten:
                    klanten[klant_name] = {"name": klant_name}

            if contact_raw and str(contact_raw).strip():
                contact_name = normalize_value(contact_raw)
                if contact_name:
                    # Normalize contact name capitalization + typo correction
                    contact_name = contact_name.strip().title()
                    contact_key = contact_name.lower()
                    contact_name = CONTACT_NORMALISATIE.get(contact_key, contact_name)
                    klant_name = normalize_klant(klant_raw) if klant_raw else normalize_klant(sheet_name)
                    contacten.append({
                        "name": contact_name,
                        "klant": klant_name,
                    })

    # Also extract klanten from aanbouwdeel sheets
    for sheet_name in AANBOUWDEEL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        klant_col = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip().lower() if c else "" for c in row]
            klant_col = next((i for i, h in enumerate(headers) if h in ("klant", "klant ")), None)
        if klant_col is None:
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not is_data_row(list(row)):
                continue
            klant_raw = row[klant_col] if len(row) > klant_col else None
            if klant_raw and str(klant_raw).strip():
                klant_name = normalize_klant(klant_raw)
                if klant_name and klant_name not in klanten:
                    klanten[klant_name] = {"name": klant_name}

    # Also from machines sheet
    if MACHINE_SHEET in wb.sheetnames:
        ws = wb[MACHINE_SHEET]
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if not is_data_row(list(row)):
                continue
            if is_header_row(list(row)):
                continue
            klant_raw = row[6] if len(row) > 6 else None  # Column G = Bedrijf
            if klant_raw and str(klant_raw).strip():
                klant_name = normalize_klant(klant_raw)
                if klant_name and klant_name not in klanten:
                    klanten[klant_name] = {"name": klant_name}

    # Deduplicate contacts
    seen_contacts = set()
    unique_contacts = []
    for c in contacten:
        key = (c["name"], c["klant"])
        if key not in seen_contacts:
            seen_contacts.add(key)
            unique_contacts.append(c)

    return list(klanten.values()), unique_contacts


def extract_machines(wb):
    """Extract carriers from the master 'machines' sheet."""
    machines = []
    if MACHINE_SHEET not in wb.sheetnames:
        print(f"  WARNING: Sheet '{MACHINE_SHEET}' not found")
        return machines

    ws = wb[MACHINE_SHEET]

    for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
        if not row or not row[0]:
            continue
        if is_header_row(list(row)):
            continue

        serienummer = normalize_value(row[2]) if len(row) > 2 else ""
        # Skip rows without a meaningful serial number
        if not serienummer or serienummer.strip() in ("EMTRC", "EMTRC "):
            continue

        sn = serienummer.replace(" ", "")  # Remove spaces in serial numbers
        klant = normalize_klant(row[6]) if len(row) > 6 else ""

        # Apply serial number corrections from client feedback
        sn_klant_key = (sn, klant)
        if sn_klant_key in SERIENUMMER_KLANT_CORRECTIES:
            sn = SERIENUMMER_KLANT_CORRECTIES[sn_klant_key]

        # Apply klant corrections from client feedback
        if sn in MACHINE_KLANT_CORRECTIES:
            klant = MACHINE_KLANT_CORRECTIES[sn]

        machine = {
            "machine": normalize_value(row[0]),
            "type": normalize_type(row[1]) if len(row) > 1 else "",
            "serienummer": sn,
            "bouwjaar": normalize_value(row[3]) if len(row) > 3 else "",
            "motor_serienummer": normalize_value(row[4]) if len(row) > 4 else "",
            "productiedatum": format_date(row[5]) if len(row) > 5 else "",
            "klant": klant,
            "status": normalize_value(row[7]) if len(row) > 7 else "",
            "serienummer_motors": normalize_value(row[8]) if len(row) > 8 else "",
            "serienummer_dmc": normalize_value(row[9]) if len(row) > 9 else "",
            "serienummer_scherm": normalize_value(row[10]) if len(row) > 10 else "",
            "serienummer_acculader": normalize_value(row[11]) if len(row) > 11 else "",
            "serienummer_htc": normalize_value(row[12]) if len(row) > 12 else "",
            "serienummer_afstandslader": normalize_value(row[13]) if len(row) > 13 else "",
            "serienummer_omvormer": normalize_value(row[14]) if len(row) > 14 else "",
        }
        machines.append(machine)

    return machines


def extract_aanbouwdelen(wb):
    """Extract aanbouwdelen from productie-overzicht sheets + klant-tabs."""
    aanbouwdelen = []

    # From productie-overzicht sheets (Dumper optie, Crane optie, etc.)
    for sheet_name in AANBOUWDEEL_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Read headers from row 1 and build column index
        headers_raw = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers_raw = [str(c).strip().lower() if c else "" for c in row]

        # Build a column index mapping normalized header names to column indices
        col_idx = {}
        for i, h in enumerate(headers_raw):
            if "aanbouwdeel" in h:
                col_idx["aanbouwdeel"] = i
            elif "serienummer" in h and "kraan" not in h:
                col_idx["serienummer"] = i
            elif "type" in h:
                col_idx["type"] = i
            elif "bouwjaar" in h:
                col_idx["bouwjaar"] = i
            elif "productie" in h:
                col_idx["productiedatum"] = i
            elif "klant" in h:
                col_idx["klant"] = i
            elif "kraan" in h:
                col_idx["kraan_nummer"] = i

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            if not is_data_row(list(row)):
                continue

            def get_col(name):
                idx = col_idx.get(name)
                if idx is not None and len(row) > idx:
                    return row[idx]
                return None

            serienummer = normalize_value(get_col("serienummer"))
            if not serienummer:
                continue  # Skip entries without serial number

            # Detect status: it's typically the last populated column
            # In Dumper optie it's col G (index 6), in Crane optie there's no explicit status header
            status = ""
            last_col = len(row) - 1
            while last_col >= 0 and (row[last_col] is None or str(row[last_col]).strip() == ""):
                last_col -= 1
            if last_col > max(col_idx.values()):
                status = normalize_value(row[last_col])

            klant_raw = get_col("klant")

            aanbouwdeel = {
                "aanbouwdeel": normalize_value(get_col("aanbouwdeel")) or normalize_value(row[0]),
                "type": normalize_type(get_col("type")),
                "serienummer": serienummer,
                "bouwjaar": normalize_value(get_col("bouwjaar")),
                "productiedatum": format_date(get_col("productiedatum")),
                "klant": normalize_klant(klant_raw),
                "status": status,
                "kraan_nummer": normalize_value(get_col("kraan_nummer")),
            }
            aanbouwdelen.append(aanbouwdeel)

    # From klant-tabs: extract aanbouwdelen (blocks after the first header repeat)
    for sheet_name in KLANT_TABS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        in_options_block = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            row_list = list(row)
            if is_header_row(row_list):
                # Check if this is an "Optie's" header (2nd+ header in sheet)
                col_b = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
                if "optie" in col_b:
                    in_options_block = True
                else:
                    in_options_block = False
                continue

            if not in_options_block:
                continue
            if not is_data_row(row_list):
                continue

            klant_raw = row[0]
            machine = normalize_value(row[1]) if len(row) > 1 else ""
            type_code = normalize_type(row[2]) if len(row) > 2 else ""
            serienummer = normalize_value(row[3]) if len(row) > 3 else ""

            if not machine:
                continue

            # Check if this aanbouwdeel already exists from productie sheets
            if serienummer:
                already_exists = any(a["serienummer"] == serienummer for a in aanbouwdelen)
                if already_exists:
                    continue

            aanbouwdeel = {
                "aanbouwdeel": machine.title(),
                "type": type_code or "",
                "serienummer": serienummer,
                "bouwjaar": normalize_value(row[4]) if len(row) > 4 else "",
                "productiedatum": "",
                "klant": normalize_klant(klant_raw) if klant_raw else normalize_klant(sheet_name),
                "status": "",
                "kraan_nummer": "",
            }

            # Add service dates from klant-tab if available
            garantie = format_date(row[6]) if len(row) > 6 else ""
            service = format_date(row[7]) if len(row) > 7 else ""
            keuring = format_date(row[8]) if len(row) > 8 else ""
            levering = format_date(row[9]) if len(row) > 9 else ""
            aanbouwdeel["garantiedatum"] = garantie
            aanbouwdeel["servicedatum"] = service
            aanbouwdeel["keuringsdatum"] = keuring
            aanbouwdeel["leveringsdatum"] = levering

            aanbouwdelen.append(aanbouwdeel)

    return aanbouwdelen


def extract_service_dates(wb):
    """Extract service/warranty dates from klant-tabs for machines (carriers)."""
    dates = []

    for sheet_name in KLANT_TABS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        in_machine_block = True  # First block is machines
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            row_list = list(row)
            if is_header_row(row_list):
                col_b = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
                in_machine_block = "optie" not in col_b and "machine" not in col_b
                if col_b in ("machine", ""):
                    in_machine_block = True
                continue

            if not in_machine_block:
                continue
            if not is_data_row(row_list):
                continue

            serienummer = normalize_value(row[3]) if len(row) > 3 else ""
            if not serienummer:
                continue

            entry = {
                "serienummer": serienummer.replace(" ", ""),
                "klant": normalize_klant(row[0]) if row[0] else normalize_klant(sheet_name),
                "contactpersoon": CONTACT_NORMALISATIE.get(normalize_value(row[5]).strip().title().lower(), normalize_value(row[5]).strip().title()) if len(row) > 5 and row[5] else "",
                "garantiedatum": format_date(row[6]) if len(row) > 6 else "",
                "servicedatum": format_date(row[7]) if len(row) > 7 else "",
                "keuringsdatum": format_date(row[8]) if len(row) > 8 else "",
                "leveringsdatum": format_date(row[9]) if len(row) > 9 else "",
                "opmerkingen": normalize_value(row[10]) if len(row) > 10 else "",
            }
            dates.append(entry)

    return dates


def extract_prijslijst(wb_prijslijst):
    """Extract price list from Prijslijst Excel."""
    items = []
    ws = wb_prijslijst.active

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        artikelnummer = normalize_value(row[0])
        omschrijving = normalize_value(row[1]) if len(row) > 1 else ""
        prijs = row[2] if len(row) > 2 else ""

        if not artikelnummer:
            continue

        # Handle "?" prices
        if str(prijs).strip() == "?":
            prijs = ""

        items.append({
            "artikelnummer": str(artikelnummer),
            "omschrijving": omschrijving,
            "prijs": str(prijs) if prijs != "" else "",
        })

    return items


# ---------------------------------------------------------------------------
# CSV Writers
# ---------------------------------------------------------------------------

def write_csv(filename, rows, fieldnames):
    """Write list of dicts to CSV."""
    filepath = CLEANED_DIR / filename
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Written: {filepath} ({len(rows)} rows)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    CLEANED_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("Emovr Migration — Step 1: Scan & Clean")
    print("=" * 60)

    # Load workbooks
    machines_file = INPUT_DIR / "Machines&Klanten Emovr.xlsx"
    prijslijst_file = INPUT_DIR / "Prijslijst Emovr 2026.xlsx"

    print(f"\nLoading {machines_file.name}...")
    wb_machines = openpyxl.load_workbook(machines_file, data_only=True)
    print(f"  Sheets: {wb_machines.sheetnames}")
    print(f"  Skipping: {SKIP_SHEETS}")

    print(f"\nLoading {prijslijst_file.name}...")
    wb_prijslijst = openpyxl.load_workbook(prijslijst_file, data_only=True)

    # Extract
    print("\n--- Extracting klanten & contactpersonen ---")
    klanten, contacten = extract_klanten_and_contacts(wb_machines)
    write_csv("klanten.csv", klanten, ["name"])
    write_csv("contactpersonen.csv", contacten, ["name", "klant"])

    print("\n--- Extracting machines (carriers) ---")
    machines = extract_machines(wb_machines)
    write_csv("machines.csv", machines, [
        "machine", "type", "serienummer", "bouwjaar", "motor_serienummer",
        "productiedatum", "klant", "status",
        "serienummer_motors", "serienummer_dmc", "serienummer_scherm",
        "serienummer_acculader", "serienummer_htc", "serienummer_afstandslader",
        "serienummer_omvormer",
    ])

    print("\n--- Extracting service dates ---")
    service_dates = extract_service_dates(wb_machines)
    write_csv("service_dates.csv", service_dates, [
        "serienummer", "klant", "contactpersoon",
        "garantiedatum", "servicedatum", "keuringsdatum", "leveringsdatum",
        "opmerkingen",
    ])

    print("\n--- Extracting aanbouwdelen ---")
    aanbouwdelen = extract_aanbouwdelen(wb_machines)
    write_csv("aanbouwdelen.csv", aanbouwdelen, [
        "aanbouwdeel", "type", "serienummer", "bouwjaar", "productiedatum",
        "klant", "status", "kraan_nummer",
        "garantiedatum", "servicedatum", "keuringsdatum", "leveringsdatum",
    ])

    print("\n--- Extracting prijslijst ---")
    prijslijst = extract_prijslijst(wb_prijslijst)
    write_csv("prijslijst.csv", prijslijst, ["artikelnummer", "omschrijving", "prijs"])

    # -----------------------------------------------------------------------
    # Consultant Report: 3-tier confidence feedback
    # -----------------------------------------------------------------------
    report = []
    report.append("")
    report.append("=" * 70)
    report.append("SCAN RAPPORT — Emovr Data Migratie")
    report.append("=" * 70)

    # --- TIER 1: CLEAR ---------------------------------------------------
    report.append("")
    report.append("## HELDER — Dit heb ik met zekerheid kunnen uitlezen")
    report.append("")

    report.append(f"Prijslijst: {len(prijslijst)} producten uit 'Prijslijst Emovr 2026.xlsx'")
    report.append("  Nette tabel met artikelnummers, omschrijvingen en prijzen.")
    for p in prijslijst:
        prijs_str = f"EUR {p['prijs']}" if p['prijs'] else "(geen prijs)"
        report.append(f"    {p['artikelnummer']}  {p['omschrijving']:30s}  {prijs_str}")

    report.append("")
    report.append(f"Klanten: {len(klanten)} unieke bedrijven gevonden")
    for k in sorted(klanten, key=lambda x: x['name']):
        report.append(f"    {k['name']}")

    report.append("")
    report.append(f"Contactpersonen: {len(contacten)} personen gekoppeld aan klanten")
    for c in sorted(contacten, key=lambda x: x['name']):
        report.append(f"    {c['name']:30s} -> {c['klant']}")

    report.append("")
    machines_with_klant = [m for m in machines if m['klant']]
    report.append(f"Machines (carriers): {len(machines_with_klant)} carriers met duidelijke klant-toewijzing")
    for m in machines_with_klant:
        report.append(f"    {m['serienummer']:20s} -> {m['klant']:30s} ({m['status']})")

    report.append("")
    aanbouwdelen_ok = [a for a in aanbouwdelen if a['serienummer'] and a['klant']]
    report.append(f"Aanbouwdelen: {len(aanbouwdelen_ok)} items met serienummer + klant")
    for a in aanbouwdelen_ok:
        report.append(f"    {a['aanbouwdeel']:15s} {a['type']:8s} {a['serienummer']:20s} -> {a['klant']}")

    # --- TIER 2: ASSUMPTIONS ---------------------------------------------
    report.append("")
    report.append("## AANNAMES — Hier was ik niet 100% zeker, dit heb ik zo ingeschat")
    report.append("")

    report.append("Klantnaam-normalisatie (spellingsvarianten samengevoegd):")
    name_fixes = [
        ("'Verwoert', 'Restauratie bedrijf Verwoerd'", "Restauratiebedrijf Verwoerd"),
        ("'SS Teknik AS'", "SS Teknikk AS"),
        ("'Sloetjes Montage'", "A Sloetjes montage"),
        ("'Marinus Slinerland' (typo)", "Marinus Slingerland"),
    ]
    for original, normalized in name_fixes:
        report.append(f"    {original:50s} -> {normalized}")

    report.append("")
    report.append("Type-normalisatie (varianten samengevoegd):")
    type_fixes = [
        ("'Trc 1.5', 'TRC1,5'", "TRC1.5"),
        ("'CR0,5'", "CR0.5"),
        ("'GL1,0'", "GL1.0"),
    ]
    for original, normalized in type_fixes:
        report.append(f"    {original:50s} -> {normalized}")

    report.append("")
    report.append("Overgeslagen sheets:")
    report.append("    'rongen systeem'      -> Bevat alleen 'materialen', geen bruikbare data")
    report.append("    'brouwer egalisatie'   -> Duplicaat van machines sheet rij 16")

    report.append("")
    report.append("Data-deduplicatie:")
    report.append("    Klant-tabs en 'machines' sheet bevatten overlappende data.")
    report.append("    Ik gebruik 'machines' als bron voor carriers, klant-tabs voor")
    report.append("    aanbouwdelen en service-datums.")

    # --- TIER 3: NEED HELP -----------------------------------------------
    report.append("")
    report.append("## HULP NODIG — Hier kom ik niet uit zonder jouw input")
    report.append("")

    need_help = []

    # Incomplete serial numbers
    incomplete_sns = [m for m in machines if m['serienummer'].endswith("00") and len(m['serienummer']) < 16]
    if incomplete_sns:
        need_help.append({
            "vraag": "Onvolledige serienummers",
            "detail": [f"  '{m['serienummer']}' (klant: {m['klant'] or '?'}) — mist laatste cijfers" for m in incomplete_sns],
            "actie": "Kun je de volledige serienummers aanleveren?",
        })

    # Machines without customer
    orphans = [m for m in machines if not m['klant']]
    if orphans:
        need_help.append({
            "vraag": f"{len(orphans)} machines zonder klant",
            "detail": [f"  {m['serienummer']:20s} status: {m['status'] or 'onbekend'}" for m in orphans],
            "actie": "Zijn dit machines in productie/voorraad? Of moeten ze aan een klant gekoppeld worden?",
        })

    # Missing prices
    missing_prices = [p for p in prijslijst if not p['prijs']]
    if missing_prices:
        need_help.append({
            "vraag": f"{len(missing_prices)} producten zonder prijs (stond als '?' in Excel)",
            "detail": [f"  {p['omschrijving']}" for p in missing_prices],
            "actie": "Wat zijn de prijzen? Of moeten deze producten overgeslagen worden?",
        })

    # Aanbouwdelen without serial
    no_serial = [a for a in aanbouwdelen if not a['serienummer']]
    if no_serial:
        need_help.append({
            "vraag": "Aanbouwdelen zonder serienummer",
            "detail": [f"  {a['aanbouwdeel']} {a['type']} (klant: {a['klant']})" for a in no_serial],
            "actie": "Is hier een serienummer voor? Of N.V.T. en gewoon als product importeren?",
        })

    # Brouwer vs Brouwer Egalisatie
    need_help.append({
        "vraag": "'Brouwer' en 'Brouwer Egalisatie' — is dit dezelfde klant?",
        "detail": ["  Brouwer Egalisatie komt voor als aparte tab en in machines sheet"],
        "actie": "Samenvoegen tot 1 klant, of 2 aparte bedrijven?",
    })

    # Service dates model question
    need_help.append({
        "vraag": "Waar moeten garantie/service/keuringsdatums heen in Odoo?",
        "detail": [
            f"  Ik heb {len(service_dates)} datum-combinaties gevonden op de klant-tabs",
            "  Opties: velden op stock.lot, apart model, of maintenance module",
        ],
        "actie": "Welke Odoo-modules zijn actief? (Inventory, Maintenance, Field Service?)",
    })

    for i, item in enumerate(need_help, 1):
        report.append(f"  {i}. {item['vraag']}")
        for d in item['detail']:
            report.append(f"     {d}")
        report.append(f"     -> {item['actie']}")
        report.append("")

    # --- SUMMARY ---------------------------------------------------------
    report.append("=" * 70)
    report.append("SAMENVATTING")
    report.append("=" * 70)
    report.append(f"  Bestanden gelezen:    2 Excel bestanden, {len(wb_machines.sheetnames)} sheets")
    report.append(f"  Klanten:              {len(klanten)}")
    report.append(f"  Contactpersonen:      {len(contacten)}")
    report.append(f"  Machines (carriers):  {len(machines)}")
    report.append(f"  Aanbouwdelen:         {len(aanbouwdelen)}")
    report.append(f"  Prijslijst items:     {len(prijslijst)}")
    report.append(f"  Service datums:       {len(service_dates)}")
    report.append(f"  Vragen voor klant:    {len(need_help)}")
    report.append(f"")
    report.append(f"  Output: {CLEANED_DIR}/")

    # Print report
    report_text = "\n".join(report)
    print(report_text)

    # Also write report to file
    report_path = CLEANED_DIR / "RAPPORT.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)
    print(f"\n  Rapport opgeslagen: {report_path}")


if __name__ == "__main__":
    main()
